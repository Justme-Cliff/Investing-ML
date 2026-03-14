# advisor/exporter.py — Export results to Book1.xlsx (6 sheets, formatted)

import os
from datetime import datetime
from typing import List, Optional

import pandas as pd

from config import FACTOR_NAMES

BOOK_PATH = "Book1.xlsx"

# Colour palette (openpyxl hex, no #)
HDR_FILL  = "1F3A5F"   # dark blue header
HDR_FONT  = "E6EDF3"   # light text
HIGH_FILL = "1A7F37"   # green — high score
MID_FILL  = "B08800"   # amber — medium
LOW_FILL  = "6E2323"   # red — low

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


class ExcelExporter:

    def export(self, top10: pd.DataFrame, macro_data: dict,
               profile, memory, allocation_df: pd.DataFrame,
               protocol_results: list = None,
               valuation_results: dict = None,
               risk_results: dict = None):
        if not _OPENPYXL:
            print("  openpyxl not installed — skipping Excel export.")
            print("  Run:  pip install openpyxl")
            return

        # Load existing workbook or create fresh
        if os.path.exists(BOOK_PATH):
            try:
                wb = load_workbook(BOOK_PATH)
            except Exception:
                wb = Workbook()
        else:
            wb = Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # Write / overwrite each sheet
        self._write_picks(wb, top10)
        self._write_allocation(wb, allocation_df, profile.portfolio_size)
        self._write_macro(wb, macro_data)
        self._write_history(wb, memory)
        self._write_track_record(wb, memory)
        if protocol_results:
            self._write_deep_analysis(wb, protocol_results, valuation_results, risk_results)

        n_sheets = len(wb.sheetnames)
        wb.save(BOOK_PATH)
        print(f"  Excel export saved → {BOOK_PATH}  ({n_sheets} sheets)")

    # ── Helpers ───────────────────────────────────────────────────────────────
    @staticmethod
    def _get_or_create_sheet(wb, name: str):
        if name in wb.sheetnames:
            ws = wb[name]
            ws.delete_rows(1, ws.max_row)   # clear content, keep sheet
        else:
            ws = wb.create_sheet(name)
        return ws

    @staticmethod
    def _header_row(ws, cols: list, row: int = 1):
        hdr_fill = PatternFill("solid", fgColor=HDR_FILL)
        hdr_font = Font(bold=True, color=HDR_FONT, size=10)
        for ci, col_name in enumerate(cols, 1):
            cell = ws.cell(row=row, column=ci, value=col_name)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = max(len(str(col_name)) + 4, 12)

    @staticmethod
    def _score_fill(score: float) -> PatternFill:
        if score >= 70:
            return PatternFill("solid", fgColor=HIGH_FILL)
        if score >= 45:
            return PatternFill("solid", fgColor=MID_FILL)
        return PatternFill("solid", fgColor=LOW_FILL)

    # ── Sheet 1: Latest Picks ─────────────────────────────────────────────────
    def _write_picks(self, wb, top10: pd.DataFrame):
        ws = self._get_or_create_sheet(wb, "Latest Picks")
        ws["A1"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A1"].font = Font(italic=True, color="888888", size=9)
        ws.row_dimensions[1].height = 14

        factor_cols = [f"{f}_score" for f in FACTOR_NAMES]
        avail_factors = [fc for fc in factor_cols if fc in top10.columns]
        factor_labels = [fc.replace("_score", "").capitalize() for fc in avail_factors]

        headers = ["Rank", "Ticker", "Sector", "Composite"] + factor_labels + ["Div%", "Price"]
        self._header_row(ws, headers, row=2)

        for ri, (_, row) in enumerate(top10.iterrows(), 3):
            vals = [
                int(row["rank"]),
                row["ticker"],
                row["sector"],
                round(float(row["composite_score"]), 1),
            ]
            for fc in avail_factors:
                vals.append(round(float(row.get(fc, 0)), 1))
            vals += [round(float(row.get("div_pct", 0)), 2),
                     round(float(row["current_price"]), 2)]

            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = Alignment(horizontal="center")
                # Colour score cells
                if ci == 4:  # composite
                    cell.fill = self._score_fill(float(val) if val else 0)
                    cell.font = Font(bold=True, color=HDR_FONT)

        # Conditional colour scale on factor columns (cols 5 → 5+n)
        if avail_factors:
            from openpyxl.formatting.rule import ColorScaleRule
            start_col = 5
            end_col   = start_col + len(avail_factors) - 1
            start_cell = f"{get_column_letter(start_col)}3"
            end_cell   = f"{get_column_letter(end_col)}{2+len(top10)}"
            rule = ColorScaleRule(
                start_type="num", start_value=0,  start_color="DA3633",
                mid_type="num",   mid_value=50,   mid_color="E3B341",
                end_type="num",   end_value=100,  end_color="3FB950",
            )
            ws.conditional_formatting.add(f"{start_cell}:{end_cell}", rule)

    # ── Sheet 2: Allocation ───────────────────────────────────────────────────
    def _write_allocation(self, wb, df: pd.DataFrame, portfolio_size: float):
        ws = self._get_or_create_sheet(wb, "Allocation")
        ws["A1"] = f"Portfolio: ${portfolio_size:,.0f}   |   Generated: {datetime.now().strftime('%Y-%m-%d')}"
        ws["A1"].font = Font(italic=True, color="888888", size=9)

        headers = ["Rank", "Ticker", "Sector", "Weight%", "$Amount", "Approx Shares", "Price"]
        self._header_row(ws, headers, row=2)

        for ri, (_, row) in enumerate(df.iterrows(), 3):
            vals = [
                int(row.get("rank", ri - 2)),
                row["ticker"],
                row["sector"],
                round(float(row.get("weight", 0)) * 100, 2),
                round(float(row.get("dollar_amount", 0)), 2),
                row.get("approx_shares", "?"),
                round(float(row["current_price"]), 2),
            ]
            for ci, val in enumerate(vals, 1):
                ws.cell(row=ri, column=ci, value=val).alignment = Alignment(horizontal="center")

    # ── Sheet 3: Macro Overview ───────────────────────────────────────────────
    def _write_macro(self, wb, macro_data: dict):
        ws = self._get_or_create_sheet(wb, "Macro Overview")
        hdr_fill = PatternFill("solid", fgColor=HDR_FILL)
        hdr_font = Font(bold=True, color=HDR_FONT)

        def row(r, label, value):
            ws.cell(r, 1, label).font = hdr_font
            ws.cell(r, 1).fill = hdr_fill
            ws.cell(r, 2, str(value))

        row(1, "Generated",     datetime.now().strftime("%Y-%m-%d %H:%M"))
        row(2, "Regime",        macro_data.get("regime", "neutral").upper())
        row(3, "VIX",           f"{macro_data.get('vix', 'N/A'):.1f}" if macro_data.get("vix") else "N/A")
        row(4, "10Y Yield",     f"{macro_data.get('yield_10y', 'N/A'):.2f}%" if macro_data.get("yield_10y") else "N/A")
        row(5, "Regime Signal", "  |  ".join(macro_data.get("regime_reasons", [])))

        ws.cell(7, 1, "Sector ETF 3-Month Returns").font = Font(bold=True)
        etf = macro_data.get("sector_etf", {})
        ri = 8
        for sector, ret in sorted(etf.items(), key=lambda x: x[1], reverse=True):
            ws.cell(ri, 1, sector)
            ws.cell(ri, 2, f"{ret:+.2f}%")
            ri += 1

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 30

    # ── Sheet 4: History ──────────────────────────────────────────────────────
    def _write_history(self, wb, memory):
        ws = self._get_or_create_sheet(wb, "History")
        headers = ["Session ID", "Date", "Risk", "Horizon", "Goal", "Tickers Picked", "# Picks"]
        self._header_row(ws, headers, row=1)

        for ri, s in enumerate(reversed(memory._data["sessions"]), 2):
            tickers = ", ".join(p["ticker"] for p in s["picks"])
            prof    = s.get("profile", {})
            vals    = [
                s.get("session_id", "?"),
                s["timestamp"][:10],
                prof.get("risk_level", "?"),
                prof.get("time_horizon", "?"),
                prof.get("goal", "?"),
                tickers,
                len(s["picks"]),
            ]
            for ci, val in enumerate(vals, 1):
                ws.cell(row=ri, column=ci, value=val)

    # ── Sheet 6: Deep Quantitative Analysis ──────────────────────────────────
    def _write_deep_analysis(self, wb, protocol_results: list,
                             valuation_results: dict = None,
                             risk_results: dict = None):
        ws = self._get_or_create_sheet(wb, "Deep Analysis")
        val_map  = valuation_results or {}
        risk_map = risk_results or {}

        ws["A1"] = f"Deep Quantitative Analysis  —  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A1"].font = Font(bold=True, color="58a6ff", size=11)
        ws.row_dimensions[1].height = 18
        ws["A2"] = "7-Gate Protocol  ·  DCF · Graham · EV/EBITDA · FCF Yield  ·  ROIC/WACC · Piotroski · Altman Z  ·  No AI APIs required"
        ws["A2"].font = Font(italic=True, color="8b949e", size=9)
        ws.row_dimensions[2].height = 14

        # ── Section 1: Gate scorecard ─────────────────────────────────────────
        gate_headers = [
            "Rank", "Ticker", "Protocol Score", "Conviction",
            "Quality", "Moat", "Health", "Valuation", "Tech Entry", "News", "Trend",
            "Pass", "Warn", "Fail",
            "Fair Value", "Entry Low", "Entry High", "Current Price", "Signal", "Methods",
        ]
        self._header_row(ws, gate_headers, row=4)

        SIGNAL_FILL = {
            "STRONG_BUY":  "1A7F37",
            "BUY":         "1F6BAA",
            "HOLD_WATCH":  "7D5A00",
            "WAIT":        "8A3700",
            "AVOID_PEAK":  "6E2323",
        }

        for ri, p in enumerate(protocol_results, 5):
            t     = p["ticker"]
            ea    = p.get("entry_analysis", {})
            gates = p.get("gates", [0] * 7)
            val   = val_map.get(t, {})

            # Prefer ValuationEngine data for fair value columns
            fv   = val.get("fair_value")   or ea.get("fair_value")
            elo  = val.get("entry_low")    or ea.get("entry_target")
            ehi  = val.get("entry_high")   or ea.get("entry_target")
            cur  = val.get("current_price") or ea.get("current_price")
            sig  = val.get("signal")       or ea.get("signal", "N/A")
            nmth = val.get("methods_count") or ea.get("num_methods", 0)

            row_data = [
                ri - 4, t,
                round(p.get("overall_score", 0), 1),
                p.get("conviction", "?"),
            ] + [round(g, 0) for g in gates[:7]] + [
                p.get("pass_count", 0),
                p.get("warn_count", 0),
                p.get("fail_count", 0),
                f"${fv:,.2f}"  if fv  else "N/A",
                f"${elo:,.2f}" if elo else "N/A",
                f"${ehi:,.2f}" if ehi else "N/A",
                f"${cur:,.2f}" if cur else "N/A",
                sig,
                nmth,
            ]
            for ci, val_ in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=val_)
                cell.alignment = Alignment(horizontal="center")
                if 5 <= ci <= 11:
                    v = float(val_) if val_ else 0
                    cell.fill = self._score_fill(v)
                    cell.font = Font(color=HDR_FONT, size=9)
                if ci == 19:   # Signal column
                    sf = SIGNAL_FILL.get(str(val_), "")
                    if sf:
                        cell.fill = PatternFill("solid", fgColor=sf)
                        cell.font = Font(bold=True, color=HDR_FONT)

        # Auto-width for gate table
        for col in ws.columns:
            max_len = 0
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 28)

        # ── Section 2: Valuation detail ───────────────────────────────────────
        row_start = len(protocol_results) + 8
        ws.cell(row_start, 1, "MULTI-METHOD VALUATION DETAIL").font = Font(bold=True, color="58a6ff", size=11)
        row_start += 1
        val_headers = [
            "Ticker", "DCF", "Graham Number", "EV/EBITDA Target", "FCF Yield Target",
            "Fair Value (median)", "Entry Low (−20%)", "Entry High (−10%)",
            "Target Price (+20%)", "Stop Loss", "Premium%", "Upside%", "R/R Ratio", "Signal",
        ]
        self._header_row(ws, val_headers, row=row_start)
        row_start += 1
        for p in protocol_results:
            t   = p["ticker"]
            val = val_map.get(t, {})
            est = val.get("estimates", {})
            row_data = [
                t,
                f"${est['dcf']:,.2f}"        if est.get("dcf")        else "N/A",
                f"${est['graham']:,.2f}"      if est.get("graham")     else "N/A",
                f"${est['ev_ebitda']:,.2f}"   if est.get("ev_ebitda")  else "N/A",
                f"${est['fcf_yield']:,.2f}"   if est.get("fcf_yield")  else "N/A",
                f"${val['fair_value']:,.2f}"  if val.get("fair_value") else "N/A",
                f"${val['entry_low']:,.2f}"   if val.get("entry_low")  else "N/A",
                f"${val['entry_high']:,.2f}"  if val.get("entry_high") else "N/A",
                f"${val['target_price']:,.2f}" if val.get("target_price") else "N/A",
                f"${val['stop_loss']:,.2f}"   if val.get("stop_loss")  else "N/A",
                f"{val['premium_pct']:+.1f}%" if val.get("premium_pct") is not None else "N/A",
                f"{val['upside_pct']:+.1f}%"  if val.get("upside_pct") is not None else "N/A",
                f"{val['rr_ratio']:.2f}:1"    if val.get("rr_ratio")  else "N/A",
                val.get("signal", "N/A"),
            ]
            for ci, v in enumerate(row_data, 1):
                cell = ws.cell(row=row_start, column=ci, value=v)
                cell.alignment = Alignment(horizontal="center")
            row_start += 1

        # ── Section 3: Risk metrics ───────────────────────────────────────────
        row_start += 2
        ws.cell(row_start, 1, "RISK & QUALITY METRICS").font = Font(bold=True, color="58a6ff", size=11)
        row_start += 1
        risk_headers = [
            "Ticker",
            "Altman Z", "Z Zone",
            "Sharpe", "Sortino",
            "Max DD%", "VaR 95% (1mo)",
            "ROIC%", "WACC%", "ROIC-WACC Spread", "Verdict",
            "Accruals Ratio", "Gross/Assets",
            "Piotroski", "/9", "Interpretation",
        ]
        self._header_row(ws, risk_headers, row=row_start)
        row_start += 1
        for p in protocol_results:
            t    = p["ticker"]
            risk = risk_map.get(t, {})
            az   = risk.get("altman_z", {})
            rw   = risk.get("roic_wacc", {})
            pf   = risk.get("piotroski", {})
            row_data = [
                t,
                f"{az['score']:.2f}"   if az.get("score") is not None else "N/A",
                az.get("zone", "N/A"),
                f"{risk['sharpe']:.2f}"  if risk.get("sharpe")  is not None else "N/A",
                f"{risk['sortino']:.2f}" if risk.get("sortino") is not None else "N/A",
                f"{risk['max_drawdown_pct']:.1f}%" if risk.get("max_drawdown_pct") is not None else "N/A",
                f"{risk['var_95_pct']:.1f}%"       if risk.get("var_95_pct")       is not None else "N/A",
                f"{rw['roic']:.1f}%"    if rw.get("roic")   is not None else "N/A",
                f"{rw['wacc']:.1f}%"    if rw.get("wacc")   is not None else "N/A",
                f"{rw['spread']:+.1f}%" if rw.get("spread") is not None else "N/A",
                rw.get("verdict", "N/A"),
                f"{risk['accruals']:.3f}"   if risk.get("accruals")    is not None else "N/A",
                f"{risk['gross_prof']:.3f}" if risk.get("gross_prof")  is not None else "N/A",
                pf.get("score", "N/A"),
                pf.get("out_of", 9),
                pf.get("interpretation", "N/A"),
            ]
            for ci, v in enumerate(row_data, 1):
                cell = ws.cell(row=row_start, column=ci, value=v)
                cell.alignment = Alignment(horizontal="center", wrap_text=(ci == len(row_data)))
                if ci == 3:   # Z zone colour
                    zone_clr = {"SAFE": "1A7F37", "GRAY": "7D5A00", "DISTRESS": "6E2323"}.get(str(v))
                    if zone_clr:
                        cell.fill = PatternFill("solid", fgColor=zone_clr)
                        cell.font = Font(bold=True, color=HDR_FONT)
            row_start += 1

        ws.freeze_panes = "A5"

    # ── Sheet 5: Track Record ─────────────────────────────────────────────────
    def _write_track_record(self, wb, memory):
        ws = self._get_or_create_sheet(wb, "Track Record")
        headers = ["Date", "Risk", "Horizon", "Avg Return%", "S&P Return%", "Alpha%", "Result"]
        self._header_row(ws, headers, row=1)

        track = memory.get_track_record()
        for ri, s in enumerate(track, 2):
            ev    = s.get("evaluation", {})
            avg_r = ev.get("avg_pick_return")
            sp_r  = ev.get("sp500_return")
            alpha = ev.get("alpha")
            result = "BEAT S&P" if (alpha or 0) > 0 else "Lagged"
            vals  = [
                s["timestamp"][:10],
                s["profile"].get("risk_level", "?"),
                s["profile"].get("time_horizon", "?"),
                round((avg_r or 0) * 100, 2),
                round((sp_r or 0) * 100, 2),
                round((alpha or 0) * 100, 2),
                result,
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if ci == 7:
                    cell.font = Font(
                        bold=True,
                        color="3FB950" if result == "BEAT S&P" else "DA3633"
                    )
